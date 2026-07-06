"""Generación del Excel diario con openpyxl, fotos incluidas."""
import json
import os

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PilImage
from sqlalchemy.orm import Session

from app.models import Cuadrilla, Reporte

AZUL = "1F3864"
ROJO_SUAVE = "FCE4E4"
VERDE_SUAVE = "E2EFDA"
GRIS = "F2F2F2"
FOTO_ANCHO_PX = 220


def _miniatura(ruta: str, destino_dir: str) -> str | None:
    """Genera una miniatura para no inflar el Excel con fotos de varios MB."""
    try:
        img = PilImage.open(ruta)
        img.thumbnail((FOTO_ANCHO_PX, FOTO_ANCHO_PX * 3))
        os.makedirs(destino_dir, exist_ok=True)
        salida = os.path.join(destino_dir, "mini_" + os.path.basename(ruta))
        img.convert("RGB").save(salida, "JPEG", quality=80)
        return salida
    except Exception:
        return None


def generar_excel(db: Session, fecha: str, salida: str = "reporte.xlsx") -> str:
    """Crea el archivo Excel del día con todos los reportes y su resumen."""
    reportes = (
        db.query(Reporte).filter(Reporte.fecha == fecha).join(Cuadrilla)
        .order_by(Cuadrilla.nombre).all()
    )
    todas = db.query(Cuadrilla).order_by(Cuadrilla.nombre).all()
    reportaron = {r.cuadrilla_id for r in reportes}

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte diario"

    # --- Encabezado ---
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"REPORTE DIARIO DE CUADRILLAS — {fecha}"
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    encabezados = ["Cuadrilla", "Hora", "Estado", "Actividades", "Novedades", "Fotos"]
    anchos = [22, 10, 14, 52, 34, 34]
    fila = 3
    for i, (h, a) in enumerate(zip(encabezados, anchos), start=1):
        celda = ws.cell(row=fila, column=i, value=h)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=AZUL)
        celda.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = a

    borde = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    mini_dir = os.path.join(os.path.dirname(os.path.abspath(salida)) or ".", "_minis")

    # --- Filas de reportes ---
    fila += 1
    for r in reportes:
        actividades = json.loads(r.actividades_json or "[]")
        texto_act = "\n".join(
            f"• {a.get('descripcion', '')}"
            + (f" ({a['cantidad']})" if a.get("cantidad") else "")
            + (f" — 📍 {a['lugar']}" if a.get("lugar") else "")
            for a in actividades
        ) or r.texto_corregido

        tardio = r.estado_horario == "TARDIO"
        valores = [
            r.cuadrilla.nombre,
            r.hora_recepcion[:5],
            "⚠️ EXTEMPORÁNEO" if tardio else "✔ A tiempo",
            texto_act,
            r.novedades or "—",
            "",
        ]
        for i, v in enumerate(valores, start=1):
            celda = ws.cell(row=fila, column=i, value=v)
            celda.border = borde
            celda.alignment = Alignment(vertical="top", wrap_text=True)
            if tardio:
                celda.fill = PatternFill("solid", fgColor=ROJO_SUAVE)
        if tardio:
            ws.cell(row=fila, column=3).font = Font(bold=True, color="9C0006")

        # Fotos en miniatura dentro de la celda F
        alto = max(60, 18 * (texto_act.count("\n") + 2))
        offset_y = 4
        for foto in r.fotos[:3]:
            mini = _miniatura(foto.ruta_local, mini_dir)
            if mini:
                img = XlsxImage(mini)
                img.anchor = f"F{fila}"
                ws.add_image(img)
                alto = max(alto, img.height * 0.75 + offset_y)
        ws.row_dimensions[fila].height = alto
        fila += 1

    # --- Resumen ---
    fila += 1
    a_tiempo = sum(1 for r in reportes if r.estado_horario == "A_TIEMPO")
    tardios = sum(1 for r in reportes if r.estado_horario == "TARDIO")
    faltantes = [c.nombre for c in todas if c.id not in reportaron]

    resumen = [
        ("RESUMEN DEL DÍA", "", AZUL, "FFFFFF"),
        (f"Reportaron a tiempo: {a_tiempo}", "", VERDE_SUAVE, None),
        (f"Reportes extemporáneos: {tardios}", "", ROJO_SUAVE if tardios else GRIS, None),
        (f"Sin reporte: {len(faltantes)}" + (f" → {', '.join(faltantes)}" if faltantes else ""),
         "", ROJO_SUAVE if faltantes else GRIS, None),
    ]
    for texto, _, color, fuente in resumen:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)
        celda = ws.cell(row=fila, column=1, value=texto)
        celda.fill = PatternFill("solid", fgColor=color)
        celda.font = Font(bold=True, color=fuente or "000000")
        fila += 1

    wb.save(salida)
    return salida
